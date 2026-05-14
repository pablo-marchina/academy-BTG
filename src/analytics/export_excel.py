from __future__ import annotations
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..db.engine import SessionLocal
from ..db.models import Oferta

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def exportar_ofertas_excel(
    ofertas: List[Dict[str, Any]] = None,
    caminho: str = "",
) -> str:
    if not caminho:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho = f"data/ofertas_{ts}.xlsx"

    if ofertas is None:
        session = SessionLocal()
        try:
            rows = session.query(Oferta).order_by(
                Oferta.score_atratividade.desc().nullslast()
            ).limit(500).all()
            ofertas = [{
                "codigo": r.codigo, "emissor": r.emissor, "produto": r.produto,
                "indexador": r.indexador, "taxa_raw": r.taxa_raw, "rating": r.rating,
                "coordenador": r.coordenador, "volume_mm": r.volume_mm,
                "score": r.score_atratividade, "spread": r.spread_curva_bps,
                "vencimento": r.vencimento, "fonte": r.fonte,
            } for r in rows]
        finally:
            session.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ofertas"

    cabecalhos = list(ofertas[0].keys()) if ofertas else []
    for col, h in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, oferta in enumerate(ofertas, 2):
        for col_idx, chave in enumerate(cabecalhos, 1):
            ws.cell(row=row_idx, column=col_idx, value=oferta.get(chave, ""))

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    Path(caminho).parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    logger.info(f"[Excel] Exportado: {caminho} ({len(ofertas)} linhas)")
    return caminho


def exportar_analise_excel(
    ofertas: List[Dict[str, Any]],
    scores: List[Dict[str, Any]],
    caminho: str = "",
) -> str:
    if not caminho:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho = f"data/analise_{ts}.xlsx"

    wb = Workbook()

    ws = wb.active
    ws.title = "Ofertas"
    cabecalhos = ["Emissor", "Produto", "Indexador", "Taxa", "Rating", "Score", "Spread", "Confianca", "Coordenador"]
    for col, h in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, (o, s) in enumerate(zip(ofertas, scores), 2):
        vals = [
            o.get("emissor", ""), o.get("produto", ""), o.get("indexador", ""),
            o.get("taxa_raw", ""), o.get("rating", ""),
            s.get("score_total", 0), s.get("spread_bps", o.get("spread_bps")),
            s.get("score_confianca", 0), o.get("coordenador", ""),
        ]
        for col_idx, v in enumerate(vals, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    ws2 = wb.create_sheet("Decomposicao")
    decomp_headers = ["Emissor", "Score", "Premio", "Rating", "Garantia", "Liquidez", "Timing", "Origem"]
    for col, h in enumerate(decomp_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, (o, s) in enumerate(zip(ofertas, scores), 2):
        d = s.get("decomposicao", {})
        vals = [
            o.get("emissor", ""), s.get("score_total", 0),
            d.get("premio_curva"), d.get("rating"), d.get("garantia"),
            d.get("liquidez"), d.get("timing"), d.get("origem"),
        ]
        for col_idx, v in enumerate(vals, 1):
            ws2.cell(row=row_idx, column=col_idx, value=v)

    Path(caminho).parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    logger.info(f"[Excel] Analise exportada: {caminho}")
    return caminho
